import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io

class SCCalculator:
    def __init__(self):
        # Streamlit page configuration
        st.set_page_config(page_title="短路电流计算器", layout="wide")
        st.title("短路电流计算器")

        # Initialize session state for storing results
        if 'result_dfs' not in st.session_state:
            st.session_state.result_dfs = {}
        if 'files_uploaded' not in st.session_state:
            st.session_state.files_uploaded = False

        # File uploader
        st.subheader("上传CSV文件")
        self.uploaded_files = st.file_uploader("选择CSV文件", type=["csv"], accept_multiple_files=True)

        # DS and DS1 inputs
        st.subheader("输入参数")
        col1, col2 = st.columns(2)
        with col1:
            self.ds_input = st.text_input("母线名 (DS, 逗号分隔):", placeholder="例如: DS1,DS2,DS3")
        with col2:
            self.ds1_input = st.text_input("显示名称 (DS1, 逗号分隔):", placeholder="例如: Name1,Name2,Name3")

        # Calculate button
        if st.button("计算"):
            self.calculate()

        # Display results
        if st.session_state.result_dfs:
            st.subheader("计算结果")
            for file_name, df in st.session_state.result_dfs.items():
                with st.expander(f"结果: {file_name}"):
                    st.dataframe(df, use_container_width=True)

        # Export button
        if st.session_state.result_dfs:
            excel_data = self.export_to_excel()
            st.download_button(
                label="导出到Excel",
                data=excel_data,
                file_name="short_circuit_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    def calculate(self):
        if not self.uploaded_files:
            st.error("请先上传CSV文件")
            return

        ds = [x.strip() for x in self.ds_input.split(',') if x.strip()]
        ds1 = [x.strip() for x in self.ds1_input.split(',') if x.strip()]

        if not ds or not ds1:
            st.error("请填写DS和DS1")
            return

        if len(ds) != len(ds1):
            st.error("DS和DS1的条目数量必须相同")
            return

        st.session_state.result_dfs.clear()
        st.session_state.files_uploaded = True

        for uploaded_file in self.uploaded_files:
            file_name = uploaded_file.name
            try:
                # Read CSV from uploaded file
                sccp = pd.read_csv(uploaded_file, encoding='gbk', index_col=False)
                required_columns = ['母线名', '故障类型']
                if not all(col in sccp.columns for col in required_columns):
                    missing = [col for col in required_columns if col not in sccp.columns]
                    st.error(f"文件 {file_name} 缺少必要列: {', '.join(missing)}")
                    return

                # Ensure fifth column exists
                if len(sccp.columns) < 5:
                    st.error(f"文件 {file_name} 列数不足，缺少短路电流数据（第5列）")
                    return

                S2 = []
                S1 = []

                # Process data
                for i in ds:
                    found = False
                    for row in sccp.itertuples():
                        if i in row.母线名:
                            found = True
                            if row.故障类型 == '单相':
                                dict_sccp = {row.母线名: row[5]}
                                S1.append(dict_sccp)
                            elif row.故障类型 == '三相':
                                dict_sccp = {row.母线名: row[5]}
                                S2.append(dict_sccp)
                    if not found:
                        st.warning(f"文件 {file_name} 中未找到母线名包含 '{i}' 的记录")

                # Check if any data was found
                if not S1 and not S2:
                    st.error(f"文件 {file_name} 未找到任何匹配的单相或三相故障数据")
                    return

                substation2 = []
                sc2 = []
                for i in S2:
                    keys_values = i.items()
                    for key, value in keys_values:
                        substation2.append(key)
                        sc2.append(value)

                SD2 = {'substation': substation2, 'sc': sc2}
                df2 = pd.DataFrame(SD2)

                substation1 = []
                sc1 = []
                for i in S1:
                    keys_values = i.items()
                    for key, value in keys_values:
                        substation1.append(key)
                        sc1.append(value)

                SD1 = {'substation': substation1, 'sc': sc1}
                df1 = pd.DataFrame(SD1)

                # Check if DataFrames are empty
                if df2.empty and df1.empty:
                    st.error(f"文件 {file_name} 处理后未生成有效数据，请检查DS输入和CSV内容")
                    return

                X1 = list(zip(ds, ds1))
                df2c = df2.copy()
                df1c = df1.copy()
                DF2 = pd.DataFrame()
                DF1 = pd.DataFrame()

                # Assign sub_name for three-phase faults
                for i in df2.index:
                    matched = False
                    for name in X1:
                        if df2.loc[i]['substation'] == name[0]:
                            df2c.at[i, 'sub_name'] = name[1]
                            matched = True
                            break
                    if not matched:
                        df2c.at[i, 'sub_name'] = df2.loc[i]['substation']

                DF2['sub_name'] = df2c['sub_name']
                DF2['sc'] = df2c['sc']

                # Assign sub_name for single-phase faults
                for i in df1.index:
                    matched = False
                    for name in X1:
                        if df1.loc[i]['substation'] == name[0]:
                            df1c.at[i, 'sub_name'] = name[1]
                            matched = True
                            break
                    if not matched:
                        df1c.at[i, 'sub_name'] = df1.loc[i]['substation']

                DF1['sub_name'] = df1c['sub_name']
                DF1['sc'] = df1c['sc']

                result_df = pd.DataFrame()
                result_df['sub_name'] = DF2['sub_name']
                result_df['sc2'] = DF2['sc']
                result_df['sc1'] = DF1['sc']

                # Handle potential NaN values
                result_df = result_df.fillna('-')

                # Round results
                result_df[['sc2', 'sc1']] = result_df[['sc2', 'sc1']].apply(pd.to_numeric, errors='coerce').round(1)

                # Store result
                st.session_state.result_dfs[file_name] = result_df

            except Exception as e:
                st.error(f"处理文件 {file_name} 时发生错误: {str(e)}")
                return

        st.success("所有文件计算完成！")

    def export_to_excel(self):
        if not st.session_state.result_dfs:
            st.error("没有可导出的结果")
            return None

        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)

        for file_name, df in st.session_state.result_dfs.items():
            ws = wb.create_sheet(title=file_name)
            ws['A1'] = file_name
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx).value = value

        wb.save(output)
        return output.getvalue()

if __name__ == "__main__":
    app = SCCalculator()
